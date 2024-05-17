import openpyxl
from datetime import date,datetime
today = date.today()

wb = openpyxl.load_workbook('E:/Office Attendance/ERP.xlsx')

def mark_attendance():
    sheet_names = wb.sheetnames

    # Display the names of all sheets
    print("List of employees: ")
    for name in sheet_names:
        print(name)
    choose_name=input("Write name you want to mark  attendance for :").title()
    if choose_name in sheet_names:
        attendance(choose_name)
    else:
        print("No")

def calculate_salary():  
    sheet_names = wb.sheetnames
  
    for name in sheet_names:
        print(name)
    choose_name=input("Write name of employee you want to calculate salary for :").title()
    choose_month=input("Enter month:")
    per_day_sal=int(input("Enter Per day salary:"))
    bonus=int("Enter bonus amount if any:")
    # Determine the month number based on the input month name
    month_number = datetime.strptime(choose_month, '%B').month
    
    # Access the worksheet for the chosen employee
    sheet = wb[choose_name]
    
    # Define column letters for relevant columns
    date_column = 'A'
    status_column = 'D'
    
    # Count the number of absences for the specified month
    absences_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming data starts from row 2
        date_value = row[0]  # Assuming date is in the first column
        status_value = row[3]  # Assuming status is in the fourth column (adjust if needed)
        
        # Check if the date is in the specified month and the status is 'Absent'
        if date_value.month == month_number and status_value == 'Absent':
            absences_count += 1
    total_salary=(30-absences_count)*per_day_sal
    
    # Print the number of absences for the specified month
    print(f"Number of absences for {choose_name} in {choose_month}: {absences_count}")
    print(f"Salary for {choose_name} in {choose_month}: {total_salary} ")

def attendance(choose_name):
    today = datetime.today().date()
    sheet = wb[choose_name]

    # Find the next empty row in column A
    next_row = sheet.max_row + 1
    status=int(input("Enter 1 for present and 0 for absent:"))
    # Add the date, day, and status to the next empty row
    sheet[f'A{next_row}'] = today
    sheet[f'B{next_row}'] = today.strftime('%B') 
    sheet[f'C{next_row}'] = today.strftime('%A')
    sheet[f'D{next_row}'] = 'Present' if status == 1 else 'Absent'

    wb.save('E:/Office Attendance/ERP.xlsx')
    # Add the date, day, and status to the next empty row
   


def menu():
    print("Today's date:", today)   
    print("\n1. Mark Attendance")
    print("2. Calculate Salary")
    choice=int(input("Enter your choice: "))
    
    if choice==1:
        mark_attendance()
    
    if choice==2:
        calculate_salary()
    
    
menu()