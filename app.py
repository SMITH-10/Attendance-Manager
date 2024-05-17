from flask import Flask, render_template, request, send_file
import openpyxl
from datetime import datetime
import pdfkit
import os

app = Flask(__name__)
wb = openpyxl.load_workbook('E:/Office Attendance/ERP.xlsx')

# Set path to the wkhtmltopdf executable
pdfkit_config = pdfkit.configuration(wkhtmltopdf='C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe')

@app.route('/')
def index():
    sheet_names = wb.sheetnames
    today = datetime.today().date()
    return render_template('index.html', sheet_names=sheet_names, today=today)

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    choose_name = request.form.get('employee_name').title()
    status = int(request.form.get('status'))
    selected_date = request.form.get('date')

    if choose_name in wb.sheetnames:
        message = attendance(choose_name, status, selected_date)
    else:
        message = "Employee not found"
    return render_template('index.html', sheet_names=wb.sheetnames, today=datetime.today().date(), message=message)

@app.route('/calculate_salary', methods=['POST'])
def calculate_salary():
    choose_name = request.form.get('employee_name').title()
    choose_month = request.form.get('month')
    per_day_sal = int(request.form.get('per_day_salary'))
    bonus = int(request.form.get('bonus'))
    advance_amount = int(request.form.get('advance_amount', 0))

    if choose_name in wb.sheetnames:
        salary, present_days = calculate_salary_logic(choose_name, choose_month, per_day_sal, bonus, advance_amount)
        if salary is not None:
            message = f"Salary for {choose_name} in {choose_month}: {salary}. Present days: {present_days}"
            generate_salary_slip(choose_name, choose_month, per_day_sal, bonus, advance_amount, present_days, salary)
            return render_template('index.html', sheet_names=wb.sheetnames, today=datetime.today().date(), message=message, slip_link=f'/salary_slip/{choose_name}_{choose_month}.pdf')
        else:
            message = "Error calculating salary."
    else:
        message = "Employee not found"
    return render_template('index.html', sheet_names=wb.sheetnames, today=datetime.today().date(), message=message)

def attendance(choose_name, status, selected_date):
    selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    sheet = wb[choose_name]
    
    for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
        if row[0] == selected_date:
            return f"Attendance already marked for {choose_name} on {selected_date}"
    
    next_row = sheet.max_row + 1
    sheet[f'A{next_row}'] = selected_date
    sheet[f'B{next_row}'] = selected_date.strftime('%B')
    sheet[f'C{next_row}'] = selected_date.strftime('%A')
    sheet[f'D{next_row}'] = 'Present' if status == 1 else 'Absent'
    wb.save('E:/Office Attendance/ERP.xlsx')

    return f"Attendance marked for {choose_name} on {selected_date}"

def calculate_salary_logic(choose_name, choose_month, per_day_sal, bonus, advance_amount):
    sheet = wb[choose_name]
    month_number = None
    
    try:
        month_number = datetime.strptime(choose_month, '%B').month
    except ValueError:
        return None, None

    present_days = 0
    total_days = 30
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            date_value = row[0]
            if date_value.month == month_number and date_value.weekday() != 6:
                present_days += 1
    
    total_salary = (present_days * per_day_sal) + bonus - advance_amount
    return total_salary, present_days

def generate_salary_slip(name, month, per_day_salary, bonus, advance, present_days, total_salary):
    today = datetime.today().strftime('%Y-%m-%d')
    firm_name = "Kasturi Sarees"
    rendered = render_template(
        'salary_slip.html', 
        name=name, 
        firm_name=firm_name,
        month=month, 
        per_day_salary=per_day_salary, 
        bonus=bonus, 
        advance=advance, 
        present_days=present_days, 
        total_salary=total_salary,
        date=today
    )
    pdf = pdfkit.from_string(rendered, False, configuration=pdfkit_config)
    pdf_path = f'salary_slips/{name}_{month}.pdf'
    with open(pdf_path, 'wb') as f:
        f.write(pdf)

@app.route('/salary_slip/<path:filename>', methods=['GET'])
def download_salary_slip(filename):
    directory = os.path.join(app.root_path, 'salary_slips')
    return send_file(os.path.join(directory, filename), as_attachment=True)

if __name__ == '__main__':
    if not os.path.exists('salary_slips'):
        os.makedirs('salary_slips')
    app.run(debug=True)
